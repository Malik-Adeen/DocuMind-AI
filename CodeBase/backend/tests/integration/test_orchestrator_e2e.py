from __future__ import annotations

import json
import uuid
from collections.abc import Sequence
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import Document, Extraction
from app.db.queries import current_extraction
from app.export.xlsx import write_workbook
from app.pipeline.llm.client import DeploymentProfile, Endpoint, LLMClient
from app.pipeline.ocr.paddle import TextRegion
from app.pipeline.orchestrator import run_and_persist

pytestmark = pytest.mark.integration

IBAN_VALID = "PK36SCBL0000001123456702"


class FakeOCR:
    def __init__(self, regions: Sequence[TextRegion]) -> None:
        self._regions = list(regions)

    def read(self, image_path: str, *, page: int = 1) -> Sequence[TextRegion]:
        return self._regions

    def page_count(self, image_path: str) -> int:
        return 1


class Spy:
    def __init__(self, response: str) -> None:
        self.response = response
        self.calls: list[str] = []

    def __call__(self, prompt: str) -> str:
        self.calls.append(prompt)
        return self.response


def regions() -> list[TextRegion]:
    return [
        TextRegion(
            text="Router hosting invoice, total 11700.00",
            confidence=0.9,
            bbox=(0.1, 0.1, 0.5, 0.2),
        )
    ]


def llm_field(value: Any, confidence: float = 0.9) -> dict[str, Any]:
    return {
        "value": value,
        "confidence": confidence,
        "verified": False,
        "source": {"origin": "llm_inferred"},
    }


def hosted_llm(response: str) -> LLMClient:
    return LLMClient(
        profile=DeploymentProfile.PROTOTYPE,
        endpoint=Endpoint.HOSTED,
        model="fake-hosted-model",
        transport=Spy(response),
    )


def clean_invoice_body() -> str:
    return json.dumps(
        {
            "document_type": {"value": "invoice", "confidence": 0.95},
            "fields": {
                "iban": llm_field(IBAN_VALID),
                "subtotal": llm_field("10000.00"),
                "tax": llm_field("1700.00"),
                "total": llm_field("11700.00"),
            },
            "line_items": [
                {
                    "description": llm_field("Managed router"),
                    "quantity": llm_field("1"),
                    "unit_price": llm_field("10000.00"),
                    "line_total": llm_field("10000.00"),
                }
            ],
        }
    )


def test_orchestrator_persists_one_extraction_row_with_gates_and_routing(
    session: Session, document_id: uuid.UUID
) -> None:
    document = session.get(Document, document_id)
    assert document is not None

    row = run_and_persist(
        session,
        document,
        ocr=FakeOCR(regions()),
        llm=hosted_llm(clean_invoice_body()),
    )

    assert row.status in {"complete", "needs_review"}
    assert row.result["gates"], "gate verdicts must be recorded on the row"
    assert row.result["pipeline_version"]["profile"] == "prototype"

    rows = session.scalars(select(Extraction).where(Extraction.document_id == document_id)).all()
    assert len(rows) == 1

    session.refresh(document)
    assert document.status == row.status


def test_rerunning_the_same_document_creates_a_new_row_not_an_update(
    session: Session, document_id: uuid.UUID
) -> None:
    document = session.get(Document, document_id)
    assert document is not None

    first = run_and_persist(
        session, document, ocr=FakeOCR(regions()), llm=hosted_llm(clean_invoice_body())
    )
    second = run_and_persist(
        session, document, ocr=FakeOCR(regions()), llm=hosted_llm(clean_invoice_body())
    )

    assert first.id != second.id

    rows = session.scalars(
        select(Extraction).where(Extraction.document_id == document_id).order_by(Extraction.seq)
    ).all()
    assert len(rows) == 2
    assert [row.id for row in rows] == [first.id, second.id]

    view = current_extraction(session, document_id)
    assert view is not None
    assert view["fields"]["iban"]["value"] == IBAN_VALID


def test_export_reads_the_orchestrator_produced_row_unmodified(
    session: Session, document_id: uuid.UUID, tmp_path: Path
) -> None:
    document = session.get(Document, document_id)
    assert document is not None

    run_and_persist(session, document, ocr=FakeOCR(regions()), llm=hosted_llm(clean_invoice_body()))

    target = tmp_path / "export.xlsx"
    write_workbook(session, document_ids=[document_id], target=target)

    assert target.exists()
    workbook = load_workbook(target)
    sheet = workbook.active
    assert sheet is not None
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header[:3] == ["document_id", "document_type", "status"]
    assert "iban" in header

    data_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    assert data_row[0] == str(document_id)
    iban_column = header.index("iban")
    assert data_row[iban_column] == IBAN_VALID
