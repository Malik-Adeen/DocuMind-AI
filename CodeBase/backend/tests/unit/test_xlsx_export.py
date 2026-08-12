from __future__ import annotations

import uuid
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from app.export import xlsx as xlsx_module
from app.schemas.extraction import field_names


def _view(**fields: dict[str, Any]) -> dict[str, Any]:
    return {
        "status": "needs_review",
        "document_type": {"value": "invoice", "confidence": 1.0},
        "fields": fields,
    }


@pytest.fixture
def document_id() -> uuid.UUID:
    return uuid.uuid4()


def _write(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, view: dict[str, Any], doc_id: uuid.UUID
) -> Path:
    monkeypatch.setattr(
        xlsx_module,
        "current_extraction",
        lambda db, requested_id: view if requested_id == doc_id else None,
    )
    target = tmp_path / "export.xlsx"
    return xlsx_module.write_workbook(db=None, document_ids=[doc_id], target=target)  # type: ignore[arg-type]


def test_header_row_matches_schema_field_order(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, document_id: uuid.UUID
) -> None:
    path = _write(monkeypatch, tmp_path, _view(), document_id)
    sheet = load_workbook(path).active
    header = [cell.value for cell in sheet[1]]
    assert header == [*xlsx_module.LEADING_COLUMNS, *field_names(), xlsx_module.UNVERIFIED_COLUMN]


def test_verified_field_value_untouched_no_styling(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, document_id: uuid.UUID
) -> None:
    view = _view(
        subtotal={
            "value": "25000.00",
            "verified": True,
            "gate": "arithmetic_reconciliation",
            "gate_error": None,
        }
    )
    path = _write(monkeypatch, tmp_path, view, document_id)
    sheet = load_workbook(path).active
    col = len(xlsx_module.LEADING_COLUMNS) + field_names().index("subtotal") + 1
    cell = sheet.cell(row=2, column=col)
    assert cell.value == "25000.00"
    assert cell.fill.fgColor.rgb in (None, "00000000")
    assert cell.comment is None


def test_unverified_field_gets_fill_and_comment_value_untouched(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, document_id: uuid.UUID
) -> None:
    view = _view(
        po_number={
            "value": "PO-2291",
            "verified": False,
            "gate": None,
            "gate_error": None,
        }
    )
    path = _write(monkeypatch, tmp_path, view, document_id)
    sheet = load_workbook(path).active
    col = len(xlsx_module.LEADING_COLUMNS) + field_names().index("po_number") + 1
    cell = sheet.cell(row=2, column=col)
    assert cell.value == "PO-2291"
    assert cell.fill.fgColor.rgb == "00FDE68A"
    assert cell.comment is not None
    assert cell.comment.text == xlsx_module.NO_GATE_REASON


def test_unverified_field_with_gate_error_uses_gate_error_as_comment(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, document_id: uuid.UUID
) -> None:
    view = _view(
        otc={
            "value": "5000.00",
            "verified": False,
            "gate": "arithmetic_reconciliation",
            "gate_error": "otc 5000.00 lacks a verbatim field label in the source text",
        }
    )
    path = _write(monkeypatch, tmp_path, view, document_id)
    sheet = load_workbook(path).active
    col = len(xlsx_module.LEADING_COLUMNS) + field_names().index("otc") + 1
    cell = sheet.cell(row=2, column=col)
    assert cell.value == "5000.00"
    assert cell.fill.fgColor.rgb == "00FDE68A"
    assert cell.comment is not None
    assert cell.comment.text == "otc 5000.00 lacks a verbatim field label in the source text"


def test_field_with_no_value_is_never_flagged(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, document_id: uuid.UUID
) -> None:
    view = _view(
        mrc={
            "value": None,
            "verified": False,
            "gate": "arithmetic_reconciliation",
            "gate_error": "mrc/otc reconciliation not computable: mrc, otc absent",
        }
    )
    path = _write(monkeypatch, tmp_path, view, document_id)
    sheet = load_workbook(path).active
    # mrc's value is None -> nothing to flag, per the "empty cell, nothing to distrust" rule.
    col = len(xlsx_module.LEADING_COLUMNS) + field_names().index("mrc") + 1
    cell = sheet.cell(row=2, column=col)
    assert cell.value is None  # openpyxl round-trips a written "" as a blank cell
    assert cell.comment is None
    summary_col = len(xlsx_module.LEADING_COLUMNS) + len(field_names()) + 1
    assert "mrc" not in (sheet.cell(row=2, column=summary_col).value or "")


def test_summary_column_lists_only_unverified_field_names(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, document_id: uuid.UUID
) -> None:
    view = _view(
        subtotal={
            "value": "25000.00",
            "verified": True,
            "gate": "arithmetic_reconciliation",
            "gate_error": None,
        },
        po_number={"value": "PO-2291", "verified": False, "gate": None, "gate_error": None},
        vendor_name={"value": "PTCL", "verified": False, "gate": None, "gate_error": None},
    )
    path = _write(monkeypatch, tmp_path, view, document_id)
    sheet = load_workbook(path).active
    summary_col = len(xlsx_module.LEADING_COLUMNS) + len(field_names()) + 1
    summary = sheet.cell(row=2, column=summary_col).value
    assert summary == "po_number, vendor_name"
