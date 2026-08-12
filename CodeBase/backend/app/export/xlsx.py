from __future__ import annotations

import uuid
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from sqlalchemy.orm import Session

from app.db.queries import current_extraction
from app.schemas.extraction import field_names

MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
LEADING_COLUMNS = ("document_id", "document_type", "status")
UNVERIFIED_COLUMN = "Unverified Fields"
UNVERIFIED_FILL = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
COMMENT_AUTHOR = "DocuMind AI"
NO_GATE_REASON = "Not covered by a deterministic gate — LLM-inferred only, not confirmed."


def write_workbook(db: Session, *, document_ids: list[uuid.UUID], target: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "extractions"
    names = field_names()
    sheet.append([*LEADING_COLUMNS, *names, UNVERIFIED_COLUMN])

    for document_id in document_ids:
        view = current_extraction(db, document_id)
        if view is None:
            continue

        reasons = {name: _unverified_reason(view, name) for name in names}
        unverified_names = [name for name, reason in reasons.items() if reason is not None]

        sheet.append(
            [
                str(document_id),
                _document_type(view),
                str(view.get("status", "")),
                *(_value(view, name) for name in names),
                ", ".join(unverified_names),
            ]
        )

        row_index = sheet.max_row
        for offset, name in enumerate(names):
            reason = reasons[name]
            if reason is None:
                continue
            cell = sheet.cell(row=row_index, column=len(LEADING_COLUMNS) + offset + 1)
            cell.fill = UNVERIFIED_FILL
            cell.comment = Comment(reason, COMMENT_AUTHOR)

    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target


def _document_type(view: dict[str, Any]) -> str:
    document_type = view.get("document_type") or {}
    return str(document_type.get("value", "unknown"))


def _value(view: dict[str, Any], name: str) -> str:
    field = (view.get("fields") or {}).get(name)
    if not field or field.get("value") is None:
        return ""
    return str(field["value"])


def _unverified_reason(view: dict[str, Any], name: str) -> str | None:
    """None means: nothing to flag — either the field has no value, or a gate verified it."""
    field = (view.get("fields") or {}).get(name)
    if not field or field.get("value") is None:
        return None
    if field.get("verified"):
        return None
    gate_error = field.get("gate_error")
    if gate_error:
        return str(gate_error)
    return NO_GATE_REASON
